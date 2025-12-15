import csv
from openpyxl import load_workbook

# ------------------ FILE PATHS ------------------
csv_file = "added_products.csv"
excel_file = "Product Master Sample.xlsx"
sheet_name = "Sheet1"

# ------------------ COLUMN MAPPING ------------------
# CSV column name : Excel column letter
COLUMN_MAPPING = {
   "Group Code": "A",
   "Item Group": "B",
   "Item Code": "F",
   "HS Code": "G",
   "Item Name": "H",
   "Description": "I",
   "Category": "J",
   "Stock Unit": "K",
   "Is Vatable Item": "O",
   "Item Type": "X",
}

START_ROW = 2   # Start writing from row 2 (keep headers)

# ------------------ LOAD EXCEL ------------------
wb = load_workbook(excel_file)
ws = wb[sheet_name]

# ------------------ READ CSV & WRITE TO EXCEL ------------------
with open(csv_file, newline="", encoding="utf-8") as f:
   reader = csv.DictReader(f)

   current_row = START_ROW
   for row in reader:
      for csv_col, excel_col in COLUMN_MAPPING.items():
         ws[f"{excel_col}{current_row}"] = row.get(csv_col)
      current_row += 1

# ------------------ SAVE ------------------
wb.save(excel_file)

print("CSV data successfully written to specific Excel cells.")
