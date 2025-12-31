import csv
from openpyxl import load_workbook

# ------------------ FILE PATHS ------------------
csv_file = "vendors.csv"
excel_file = "Vendor Master Sample.xlsx"
sheet_name = "VendorMaster"

# ------------------ COLUMN MAPPING ------------------
# CSV column name : Excel column letter
COLUMN_MAPPING = {
   "MainGroup": "A",
   "ACNAME": "E",
   "Address": "F",
   "VATNO": "M",
   "PARTYTYPE": "P",
}

START_ROW = 2

# ------------------ LOAD EXISTING EXCEL ------------------
wb = load_workbook(excel_file)
ws = wb[sheet_name]

# ------------------ WRITE CSV DATA TO SPECIFIC COLUMNS ------------------
with open(csv_file, newline="", encoding="utf-8") as f:
   reader = csv.DictReader(f)

   current_row = START_ROW
   for csv_row in reader:
      for csv_col, excel_col in COLUMN_MAPPING.items():
         ws[f"{excel_col}{current_row}"] = csv_row.get(csv_col, "")
      current_row += 1

# ------------------ SAVE BACK TO SAME FILE ------------------
wb.save(excel_file)

print("Data written successfully to existing Excel file.")
